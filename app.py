from flask import Flask, render_template, request, jsonify, redirect, session, send_file
import os, io
from datetime import datetime
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'bodega-aseo-2025-secret')
DATABASE_URL = os.environ.get('DATABASE_URL', '')

CATEGORIAS = ['Papel','Bolsas','Líquidos Limpieza','Desinfectantes','Paños','Guantes','Varios']
EDIFICIOS  = ['Básica','Media','Parvularia']
MESES      = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
               'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

def get_db():
    if DATABASE_URL:
        import psycopg2, psycopg2.extras
        return psycopg2.connect(DATABASE_URL), 'pg'
    import sqlite3
    conn = sqlite3.connect(os.path.join(os.path.dirname(os.path.abspath(__file__)),'bodega.db'))
    conn.row_factory = sqlite3.Row
    return conn, 'sqlite'

def db_fetchall(sql, params=()):
    conn, mode = get_db()
    if mode == 'pg':
        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql.replace('?','%s'), params)
    else:
        cur = conn.cursor()
        cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def db_fetchone(sql, params=()):
    conn, mode = get_db()
    if mode == 'pg':
        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql.replace('?','%s'), params)
    else:
        cur = conn.cursor()
        cur.execute(sql, params)
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None

def db_run(sql, params=()):
    conn, mode = get_db()
    cur = conn.cursor()
    cur.execute(sql.replace('?','%s') if mode=='pg' else sql, params)
    conn.commit()
    conn.close()

def init_db():
    conn, mode = get_db()
    cur = conn.cursor()
    PK = 'SERIAL PRIMARY KEY' if mode=='pg' else 'INTEGER PRIMARY KEY AUTOINCREMENT'
    TS = 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP' if mode=='pg' else 'TEXT DEFAULT CURRENT_TIMESTAMP'
    cur.execute(f'''CREATE TABLE IF NOT EXISTS productos (
        id {PK}, nombre TEXT NOT NULL, categoria TEXT,
        unidad TEXT DEFAULT 'unidades', stock_actual INTEGER DEFAULT 0,
        stock_minimo INTEGER DEFAULT 0, activo BOOLEAN DEFAULT TRUE
    )''')
    cur.execute(f'''CREATE TABLE IF NOT EXISTS movimientos (
        id {PK}, producto_id INTEGER NOT NULL,
        tipo TEXT NOT NULL, cantidad INTEGER NOT NULL,
        edificio TEXT, usuario TEXT, observacion TEXT,
        fecha TEXT, created_at {TS}
    )''')
    cur.execute(f'''CREATE TABLE IF NOT EXISTS usuarios (
        id {PK}, nombre TEXT, email TEXT UNIQUE,
        password TEXT, rol TEXT DEFAULT 'edificio', edificio TEXT
    )''')
    # Usuario unico compartido
    if mode == 'pg':
        cur.execute("INSERT INTO usuarios (nombre,email,password,rol,edificio) VALUES (%s,%s,%s,%s,%s) ON CONFLICT (email) DO NOTHING",
            ('Bodega', os.environ.get('ADMIN_EMAIL','bodega@colegio.cl'),
             os.environ.get('ADMIN_PASS','bodega2025'), 'admin', ''))
    else:
        cur.execute("INSERT OR IGNORE INTO usuarios (nombre,email,password,rol,edificio) VALUES (?,?,?,?,?)",
            ('Bodega', os.environ.get('ADMIN_EMAIL','bodega@colegio.cl'),
             os.environ.get('ADMIN_PASS','bodega2025'), 'admin', ''))
    conn.commit()
    conn.close()

with app.app_context():
    init_db()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session: return redirect('/login')
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session: return redirect('/login')
        if session.get('rol') != 'admin': return jsonify({'error':'Sin permisos'}), 403
        return f(*args, **kwargs)
    return decorated

# ── AUTH ─────────────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET','POST'])
def login():
    error = ''
    if request.method == 'POST':
        email    = request.form.get('email','').strip()
        password = request.form.get('password','')
        u = db_fetchone("SELECT * FROM usuarios WHERE email=? AND password=?", (email, password))
        if u:
            session['user']     = u['nombre']
            session['rol']      = u['rol']
            session['email']    = u['email']
            session['edificio'] = u.get('edificio','')
            return redirect('/')
        error = 'Email o contraseña incorrectos'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

# ── MAIN ─────────────────────────────────────────────────────────────────────
@app.route('/')
@login_required
def index():
    return render_template('index.html',
        user=session['user'], rol=session['rol'],
        edificio=session.get('edificio',''),
        categorias=CATEGORIAS, edificios=EDIFICIOS)

# ── PRODUCTOS ─────────────────────────────────────────────────────────────────
@app.route('/api/productos', methods=['GET'])
@login_required
def get_productos():
    cat = request.args.get('categoria','')
    q   = request.args.get('q','')
    sql = "SELECT * FROM productos WHERE activo=TRUE"
    params = []
    if cat:
        sql += " AND categoria=?"; params.append(cat)
    if q:
        sql += " AND nombre LIKE ?"; params.append(f'%{q}%')
    sql += " ORDER BY categoria, nombre"
    return jsonify(db_fetchall(sql, params))

@app.route('/api/productos/<int:pid>', methods=['GET'])
@login_required
def get_producto(pid):
    p = db_fetchone("SELECT * FROM productos WHERE id=?", (pid,))
    if not p: return jsonify({'error':'No encontrado'}), 404
    movs = db_fetchall(
        "SELECT * FROM movimientos WHERE producto_id=? ORDER BY created_at DESC LIMIT 50", (pid,))
    return jsonify({'producto': p, 'movimientos': movs})

@app.route('/api/productos', methods=['POST'])
@admin_required
def crear_producto():
    d = request.json
    conn, mode = get_db()
    cur = conn.cursor()
    if mode == 'pg':
        cur.execute(
            "INSERT INTO productos (nombre,categoria,unidad,stock_actual,stock_minimo) VALUES (%s,%s,%s,%s,%s) RETURNING id",
            (d['nombre'], d.get('categoria','Varios'), d.get('unidad','unidades'),
             int(d.get('stock_actual',0)), int(d.get('stock_minimo',0))))
        pid = cur.fetchone()[0]
    else:
        cur.execute(
            "INSERT INTO productos (nombre,categoria,unidad,stock_actual,stock_minimo) VALUES (?,?,?,?,?)",
            (d['nombre'], d.get('categoria','Varios'), d.get('unidad','unidades'),
             int(d.get('stock_actual',0)), int(d.get('stock_minimo',0))))
        pid = cur.lastrowid
    conn.commit(); conn.close()
    return jsonify({'ok':True,'id':pid})

@app.route('/api/productos/<int:pid>', methods=['PUT'])
@admin_required
def editar_producto(pid):
    d = request.json
    campos = ['nombre','categoria','unidad','stock_minimo']
    updates = {c:d[c] for c in campos if c in d}
    if updates:
        sets = ', '.join(f"{c}=?" for c in updates)
        db_run(f"UPDATE productos SET {sets} WHERE id=?", list(updates.values())+[pid])
    return jsonify({'ok':True})

@app.route('/api/productos/<int:pid>', methods=['DELETE'])
@admin_required
def eliminar_producto(pid):
    db_run("UPDATE productos SET activo=FALSE WHERE id=?", (pid,))
    return jsonify({'ok':True})

# ── MOVIMIENTOS ───────────────────────────────────────────────────────────────
@app.route('/api/movimientos', methods=['POST'])
@login_required
def registrar_movimiento():
    d       = request.json
    pid     = int(d['producto_id'])
    tipo    = d['tipo']  # 'entrada' o 'salida'
    cant    = int(d['cantidad'])
    edificio= d.get('edificio', session.get('edificio',''))
    obs     = d.get('observacion','')
    fecha   = d.get('fecha', datetime.now().strftime('%d-%m-%Y'))
    usuario = session['user']

    # Verificar stock suficiente en salida
    if tipo == 'salida':
        p = db_fetchone("SELECT stock_actual FROM productos WHERE id=?", (pid,))
        if not p: return jsonify({'error':'Producto no encontrado'}), 404
        if p['stock_actual'] < cant:
            return jsonify({'error':f'Stock insuficiente. Disponible: {p["stock_actual"]}'}), 400

    conn, mode = get_db()
    cur = conn.cursor()
    delta = cant if tipo == 'entrada' else -cant
    if mode == 'pg':
        cur.execute(
            "INSERT INTO movimientos (producto_id,tipo,cantidad,edificio,usuario,observacion,fecha) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (pid, tipo, cant, edificio, usuario, obs, fecha))
        cur.execute("UPDATE productos SET stock_actual = stock_actual + %s WHERE id=%s", (delta, pid))
    else:
        cur.execute(
            "INSERT INTO movimientos (producto_id,tipo,cantidad,edificio,usuario,observacion,fecha) VALUES (?,?,?,?,?,?,?)",
            (pid, tipo, cant, edificio, usuario, obs, fecha))
        cur.execute("UPDATE productos SET stock_actual = stock_actual + ? WHERE id=?", (delta, pid))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/api/movimientos', methods=['GET'])
@login_required
def get_movimientos():
    tipo     = request.args.get('tipo','')
    edificio = request.args.get('edificio','')
    mes      = request.args.get('mes','')
    pid      = request.args.get('producto_id','')
    sql = '''SELECT m.*, p.nombre as producto_nombre, p.categoria, p.unidad
             FROM movimientos m JOIN productos p ON m.producto_id=p.id WHERE 1=1'''
    params = []
    if tipo:     sql += " AND m.tipo=?";        params.append(tipo)
    if edificio: sql += " AND m.edificio=?";    params.append(edificio)
    if pid:      sql += " AND m.producto_id=?"; params.append(int(pid))
    if mes:      sql += " AND m.fecha LIKE ?";  params.append(f'%-{mes}-%')
    sql += " ORDER BY m.created_at DESC LIMIT 200"
    return jsonify(db_fetchall(sql, params))

# ── STATS / DASHBOARD ─────────────────────────────────────────────────────────
@app.route('/api/stats')
@login_required
def stats():
    total_prod  = db_fetchone("SELECT COUNT(*) as n FROM productos WHERE activo=TRUE")['n']
    bajo_minimo = db_fetchone(
        "SELECT COUNT(*) as n FROM productos WHERE activo=TRUE AND stock_actual <= stock_minimo AND stock_minimo > 0")['n']
    total_ent   = db_fetchone(
        "SELECT COALESCE(SUM(cantidad),0) as s FROM movimientos WHERE tipo='entrada'")['s']
    total_sal   = db_fetchone(
        "SELECT COALESCE(SUM(cantidad),0) as s FROM movimientos WHERE tipo='salida'")['s']
    return jsonify({'total_productos':total_prod,'bajo_minimo':bajo_minimo,
                    'total_entradas':total_ent,'total_salidas':total_sal})

@app.route('/api/kpis')
@login_required
def kpis():
    # Stock actual por categoria
    por_cat = db_fetchall(
        "SELECT categoria, COUNT(*) as productos, SUM(stock_actual) as stock_total FROM productos WHERE activo=TRUE GROUP BY categoria ORDER BY stock_total DESC")
    # Alertas stock minimo
    alertas = db_fetchall(
        "SELECT * FROM productos WHERE activo=TRUE AND stock_actual <= stock_minimo AND stock_minimo > 0 ORDER BY stock_actual ASC")
    # Consumo por edificio (salidas)
    por_edificio = db_fetchall(
        "SELECT edificio, SUM(cantidad) as total FROM movimientos WHERE tipo='salida' AND edificio!='' GROUP BY edificio ORDER BY total DESC")
    # Consumo mensual (ultimos 6 meses)
    consumo_mes = db_fetchall(
        """SELECT fecha, SUM(CASE WHEN tipo='entrada' THEN cantidad ELSE 0 END) as entradas,
                  SUM(CASE WHEN tipo='salida' THEN cantidad ELSE 0 END) as salidas
           FROM movimientos GROUP BY fecha ORDER BY fecha DESC LIMIT 100""")
    # Agrupar por mes
    mes_dict = {}
    for r in consumo_mes:
        f = str(r.get('fecha',''))
        if len(f) >= 7:
            parts = f.split('-')
            if len(parts) == 3:
                mes_key = f"{parts[1]}-{parts[2]}" if len(parts[2])==4 else f"{parts[2]}-{parts[1]}"
                if mes_key not in mes_dict:
                    mes_dict[mes_key] = {'entradas':0,'salidas':0}
                mes_dict[mes_key]['entradas'] += int(r['entradas'] or 0)
                mes_dict[mes_key]['salidas']  += int(r['salidas'] or 0)
    consumo_mensual = [{'mes':k,'entradas':v['entradas'],'salidas':v['salidas']}
                       for k,v in sorted(mes_dict.items())[-12:]]
    # Productos mas consumidos
    mas_consumidos = db_fetchall(
        """SELECT p.nombre, p.categoria, p.unidad, p.stock_actual,
                  COALESCE(SUM(CASE WHEN m.tipo='salida' THEN m.cantidad ELSE 0 END),0) as total_salidas
           FROM productos p LEFT JOIN movimientos m ON p.id=m.producto_id
           WHERE p.activo=TRUE GROUP BY p.id, p.nombre, p.categoria, p.unidad, p.stock_actual
           ORDER BY total_salidas DESC LIMIT 10""")
    # Consumo por edificio por producto (top)
    consumo_ed_prod = db_fetchall(
        """SELECT m.edificio, p.nombre, SUM(m.cantidad) as total
           FROM movimientos m JOIN productos p ON m.producto_id=p.id
           WHERE m.tipo='salida' AND m.edificio!=''
           GROUP BY m.edificio, p.nombre ORDER BY total DESC LIMIT 30""")
    return jsonify({
        'por_categoria':    por_cat,
        'alertas':          alertas,
        'por_edificio':     por_edificio,
        'consumo_mensual':  consumo_mensual,
        'mas_consumidos':   mas_consumidos,
        'consumo_ed_prod':  consumo_ed_prod,
    })

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html', user=session['user'], rol=session['rol'])

# ── EXPORT ────────────────────────────────────────────────────────────────────
@app.route('/api/export/stock')
@login_required
def export_stock():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    rows = db_fetchall("SELECT * FROM productos WHERE activo=TRUE ORDER BY categoria, nombre")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="Stock Actual"
    headers = ['ID','Nombre','Categoría','Unidad','Stock Actual','Stock Mínimo','Estado']
    for col,h in enumerate(headers,1):
        cell = ws.cell(row=1,column=col,value=h)
        cell.font = Font(bold=True,color='FFFFFF')
        cell.fill = PatternFill('solid',start_color='1F3864',fgColor='1F3864')
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width = 18
    for ri,r in enumerate(rows,2):
        estado = 'BAJO MÍNIMO' if r['stock_actual'] <= r['stock_minimo'] and r['stock_minimo']>0 else 'OK'
        for col,val in enumerate([r['id'],r['nombre'],r['categoria'],r['unidad'],r['stock_actual'],r['stock_minimo'],estado],1):
            ws.cell(row=ri,column=col,value=val)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,download_name='stock_bodega.xlsx',as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/export/movimientos')
@login_required
def export_movimientos():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    rows = db_fetchall(
        """SELECT m.fecha, m.tipo, p.nombre, p.unidad, m.cantidad,
                  m.edificio, m.usuario, m.observacion
           FROM movimientos m JOIN productos p ON m.producto_id=p.id
           ORDER BY m.created_at DESC""")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="Movimientos"
    headers = ['Fecha','Tipo','Producto','Unidad','Cantidad','Edificio','Usuario','Observación']
    for col,h in enumerate(headers,1):
        cell = ws.cell(row=1,column=col,value=h)
        cell.font = Font(bold=True,color='FFFFFF')
        cell.fill = PatternFill('solid',start_color='1F3864',fgColor='1F3864')
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width = 18
    for ri,r in enumerate(rows,2):
        for col,key in enumerate(['fecha','tipo','nombre','unidad','cantidad','edificio','usuario','observacion'],1):
            ws.cell(row=ri,column=col,value=r.get(key,''))
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,download_name='movimientos_bodega.xlsx',as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── USUARIOS ──────────────────────────────────────────────────────────────────



@app.route('/api/importar', methods=['POST'])
@login_required
def importar_excel():
    if 'archivo' not in request.files:
        return jsonify({'error':'No se envió archivo'}), 400
    file = request.files['archivo']
    if not file.filename.lower().endswith(('.xlsx','.xls')):
        return jsonify({'error':'Solo se aceptan archivos Excel (.xlsx)'}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        # Buscar hoja PRODUCTOS
        ws = None
        for name in wb.sheetnames:
            if 'PROD' in name.upper():
                ws = wb[name]; break
        if ws is None: ws = wb.active

        # Columnas fijas: A=Nombre, B=Categoria, C=Unidad, D=Stock actual, E=Stock minimo
        creados = 0
        errores = []
        for row_num in range(4, ws.max_row + 1):
            def gv(col):
                v = ws.cell(row=row_num, column=col).value
                return str(v).strip() if v is not None else ''
            nombre   = gv(1)
            categoria= gv(2)
            unidad   = gv(3)
            if not nombre or not categoria: continue
            try:
                stock   = int(float(gv(4))) if gv(4) else 0
                minimo  = int(float(gv(5))) if gv(5) else 0
                conn2, mode2 = get_db()
                cur2 = conn2.cursor()
                if mode2 == 'pg':
                    cur2.execute(
                        "INSERT INTO productos (nombre,categoria,unidad,stock_actual,stock_minimo) VALUES (%s,%s,%s,%s,%s)",
                        (nombre, categoria, unidad or 'unidades', stock, minimo))
                else:
                    cur2.execute(
                        "INSERT INTO productos (nombre,categoria,unidad,stock_actual,stock_minimo) VALUES (?,?,?,?,?)",
                        (nombre, categoria, unidad or 'unidades', stock, minimo))
                conn2.commit(); conn2.close()
                creados += 1
            except Exception as e:
                errores.append(f"Fila {row_num}: {str(e)}")
        return jsonify({'ok':True,'creados':creados,'errores':errores})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/kpis/filtrado')
@login_required
def kpis_filtrado():
    fecha_desde = request.args.get('desde','')
    fecha_hasta = request.args.get('hasta','')
    mes         = request.args.get('mes','')  # formato MM-YYYY

    # Construir filtro de fechas
    filtro_sql = "WHERE tipo='salida'"
    params = []
    if mes:
        filtro_sql += " AND fecha LIKE ?"
        params.append(f"%-{mes}")
    elif fecha_desde and fecha_hasta:
        filtro_sql += " AND fecha >= ? AND fecha <= ?"
        params += [fecha_desde, fecha_hasta]
    elif fecha_desde:
        filtro_sql += " AND fecha >= ?"
        params.append(fecha_desde)

    # Total entradas y salidas en periodo
    filtro_ent = filtro_sql.replace("tipo='salida'","tipo='entrada'")
    total_sal = db_fetchone(
        f"SELECT COALESCE(SUM(cantidad),0) as s FROM movimientos {filtro_sql}", params)
    total_ent = db_fetchone(
        f"SELECT COALESCE(SUM(cantidad),0) as s FROM movimientos {filtro_ent}", params)

    # Salidas por edificio en periodo
    por_edificio = db_fetchall(
        f"""SELECT edificio, SUM(cantidad) as total
            FROM movimientos {filtro_sql} AND edificio!=''
            GROUP BY edificio ORDER BY total DESC""", params)

    # Salidas por producto en periodo
    por_producto = db_fetchall(
        f"""SELECT p.nombre, p.categoria, p.unidad, SUM(m.cantidad) as total
            FROM movimientos m JOIN productos p ON m.producto_id=p.id
            {filtro_sql}
            GROUP BY p.id, p.nombre, p.categoria, p.unidad
            ORDER BY total DESC LIMIT 15""", params)

    # Movimientos detalle
    detalle = db_fetchall(
        f"""SELECT m.fecha, m.tipo, p.nombre, p.unidad, m.cantidad, m.edificio, m.usuario
            FROM movimientos m JOIN productos p ON m.producto_id=p.id
            {filtro_sql}
            ORDER BY m.created_at DESC LIMIT 100""", params)

    return jsonify({
        'total_salidas':  total_sal['s'] if total_sal else 0,
        'total_entradas': total_ent['s'] if total_ent else 0,
        'por_edificio':   por_edificio,
        'por_producto':   por_producto,
        'detalle':        detalle,
    })



@app.route('/api/movimientos/historial')
@login_required
def get_historial():
    tipo     = request.args.get('tipo','')
    edificio = request.args.get('edificio','')
    cat      = request.args.get('categoria_filtro','')
    q        = request.args.get('q','')
    desde    = request.args.get('desde','')
    hasta    = request.args.get('hasta','')

    sql = """SELECT m.id, m.fecha, m.tipo, m.cantidad, m.edificio,
                    m.usuario, m.observacion, m.producto_id,
                    p.nombre as producto_nombre, p.categoria, p.unidad
             FROM movimientos m JOIN productos p ON m.producto_id=p.id
             WHERE 1=1"""
    params = []
    if tipo:     sql += " AND m.tipo=?";          params.append(tipo)
    if edificio: sql += " AND m.edificio=?";      params.append(edificio)
    if cat:      sql += " AND p.categoria=?";     params.append(cat)
    if q:        sql += " AND p.nombre LIKE ?";   params.append(f'%{q}%')
    if desde:
        # desde viene como YYYY-MM-DD, convertir a DD-MM-YYYY para comparar
        parts = desde.split('-')
        if len(parts)==3:
            desde_fmt = f"{parts[2]}-{parts[1]}-{parts[0]}"
            sql += " AND m.fecha >= ?"; params.append(desde_fmt)
    if hasta:
        parts = hasta.split('-')
        if len(parts)==3:
            hasta_fmt = f"{parts[2]}-{parts[1]}-{parts[0]}"
            sql += " AND m.fecha <= ?"; params.append(hasta_fmt)
    sql += " ORDER BY m.created_at DESC LIMIT 500"
    return jsonify(db_fetchall(sql, params))

@app.route('/movimientos')
@login_required
def movimientos_page():
    return render_template('movimientos.html',
        user=session['user'], rol=session['rol'],
        categorias=CATEGORIAS, edificios=EDIFICIOS)

@app.route('/api/movimientos/<int:mid>', methods=['PUT'])
@login_required
def editar_movimiento(mid):
    d = request.json
    # Obtener movimiento original
    mov = db_fetchone("SELECT * FROM movimientos WHERE id=?", (mid,))
    if not mov: return jsonify({'error':'No encontrado'}), 404

    nueva_cant = int(d.get('cantidad', mov['cantidad']))
    nuevo_tipo = d.get('tipo', mov['tipo'])
    dif = nueva_cant - mov['cantidad']

    conn2, mode2 = get_db()
    cur2 = conn2.cursor()
    if mode2 == 'pg':
        cur2.execute(
            "UPDATE movimientos SET cantidad=%s,edificio=%s,observacion=%s,fecha=%s WHERE id=%s",
            (nueva_cant, d.get('edificio', mov['edificio']),
             d.get('observacion', mov['observacion']),
             d.get('fecha', mov['fecha']), mid))
        # Ajustar stock: si era salida, revertir diferencia
        if mov['tipo'] == 'salida':
            cur2.execute("UPDATE productos SET stock_actual=stock_actual-%s WHERE id=%s", (dif, mov['producto_id']))
        else:
            cur2.execute("UPDATE productos SET stock_actual=stock_actual+%s WHERE id=%s", (dif, mov['producto_id']))
    else:
        cur2.execute(
            "UPDATE movimientos SET cantidad=?,edificio=?,observacion=?,fecha=? WHERE id=?",
            (nueva_cant, d.get('edificio', mov['edificio']),
             d.get('observacion', mov['observacion']),
             d.get('fecha', mov['fecha']), mid))
        if mov['tipo'] == 'salida':
            cur2.execute("UPDATE productos SET stock_actual=stock_actual-? WHERE id=?", (dif, mov['producto_id']))
        else:
            cur2.execute("UPDATE productos SET stock_actual=stock_actual+? WHERE id=?", (dif, mov['producto_id']))
    conn2.commit()
    conn2.close()
    return jsonify({'ok': True})

@app.route('/api/movimientos/<int:mid>', methods=['DELETE'])
@login_required
def eliminar_movimiento(mid):
    mov = db_fetchone("SELECT * FROM movimientos WHERE id=?", (mid,))
    if not mov: return jsonify({'error':'No encontrado'}), 404
    # Revertir stock
    conn2, mode2 = get_db()
    cur2 = conn2.cursor()
    delta = mov['cantidad'] if mov['tipo'] == 'salida' else -mov['cantidad']
    if mode2 == 'pg':
        cur2.execute("UPDATE productos SET stock_actual=stock_actual+%s WHERE id=%s", (delta, mov['producto_id']))
        cur2.execute("DELETE FROM movimientos WHERE id=%s", (mid,))
    else:
        cur2.execute("UPDATE productos SET stock_actual=stock_actual+? WHERE id=?", (delta, mov['producto_id']))
        cur2.execute("DELETE FROM movimientos WHERE id=?", (mid,))
    conn2.commit()
    conn2.close()
    return jsonify({'ok': True})


@app.route('/api/importar/movimientos', methods=['POST'])
@login_required
def importar_movimientos():
    if 'archivo' not in request.files:
        return jsonify({'error':'No se envió archivo'}), 400
    file = request.files['archivo']
    if not file.filename.lower().endswith(('.xlsx','.xls')):
        return jsonify({'error':'Solo se aceptan archivos Excel (.xlsx)'}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = None
        for name in wb.sheetnames:
            if 'MOV' in name.upper():
                ws = wb[name]; break
        if ws is None: ws = wb.active

        # Columnas: A=Producto, B=Tipo, C=Cantidad, D=Edificio, E=Fecha, F=Observacion
        creados = 0
        errores = []
        for row_num in range(4, ws.max_row + 1):
            def gv(col):
                v = ws.cell(row=row_num, column=col).value
                return str(v).strip() if v is not None else ''

            nombre   = gv(1)
            tipo     = gv(2).lower()
            cant_raw = gv(3)
            edificio = gv(4)
            fecha    = gv(5)
            obs      = gv(6)

            if not nombre or not tipo or not cant_raw: continue
            if tipo not in ('entrada','salida'): continue

            try:
                cant = int(float(cant_raw))
                if cant <= 0: continue

                # Buscar producto por nombre
                prod = db_fetchone(
                    "SELECT id, stock_actual FROM productos WHERE LOWER(nombre)=LOWER(?)", (nombre,))
                if not prod:
                    errores.append(f"Fila {row_num}: producto '{nombre}' no encontrado")
                    continue

                # Verificar stock en salida
                if tipo == 'salida' and prod['stock_actual'] < cant:
                    errores.append(f"Fila {row_num}: stock insuficiente para '{nombre}' (hay {prod['stock_actual']})")
                    continue

                conn2, mode2 = get_db()
                cur2 = conn2.cursor()
                delta = cant if tipo == 'entrada' else -cant
                usu = session['user']
                if mode2 == 'pg':
                    cur2.execute(
                        "INSERT INTO movimientos (producto_id,tipo,cantidad,edificio,usuario,observacion,fecha) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                        (prod['id'], tipo, cant, edificio, usu, obs, fecha))
                    cur2.execute(
                        "UPDATE productos SET stock_actual=stock_actual+%s WHERE id=%s",
                        (delta, prod['id']))
                else:
                    cur2.execute(
                        "INSERT INTO movimientos (producto_id,tipo,cantidad,edificio,usuario,observacion,fecha) VALUES (?,?,?,?,?,?,?)",
                        (prod['id'], tipo, cant, edificio, usu, obs, fecha))
                    cur2.execute(
                        "UPDATE productos SET stock_actual=stock_actual+? WHERE id=?",
                        (delta, prod['id']))
                conn2.commit()
                conn2.close()
                creados += 1
            except Exception as e:
                errores.append(f"Fila {row_num}: {str(e)}")

        return jsonify({'ok':True,'creados':creados,'errores':errores})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

if __name__=='__main__':
    init_db()
    app.run(debug=False,host='0.0.0.0',port=int(os.environ.get('PORT',5000)))
